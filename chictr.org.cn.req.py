import random
import requests
from lxml import etree
import ddddocr
import openpyxl
import logging

logging.basicConfig(filename='logging.log', level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')


def getList(page, verifycode):
    # 获取列表页的url
    api = 'http://www.chictr.org.cn/searchproj.aspx'
    params = {
        'minstudyexecutetime': '2020-07-01',
        'province': '北京',
        'btngo': 'btn',
        'page': page,
        'verifycode': verifycode,
    }
    try:
        res = requests.get(api, headers=headers, params=params)
        e = etree.HTML(res.text)
        link_list = e.xpath('//table[@class="table_list"]/tbody/tr/td[3]/p/a/@href')
        for link in link_list:
            link = 'http://www.chictr.org.cn/' + link
            f.write(str(page) + ',' + link + '\n')
            print(page, link)
        if link_list:
            return 'ok'
        else:
            return None
    except Exception as e:
        print(e)


def getVerifyimagepage():
    # 获取验证码图片
    api = 'http://www.chictr.org.cn/Tools/verifyimagepage.aspx'
    t = random.uniform(0, 1)
    paramas = {
        'textcolor': 2,
        'bgcolor': 'F4F4F4',
        'ut': 1,
        'time': t,
    }
    try:
        res = requests.get(api, headers=headers, params=paramas)
        with open('./images/img.jpg', 'wb') as img:
            img.write(res.content)
    except Exception as e:
        print(e)


def imgRecognition(img):
    # 识别验证码
    try:
        ocr = ddddocr.DdddOcr()
        with open(img, 'rb') as f:
            img_bytes = f.read()
        res = ocr.classification(img_bytes)
        return res
    except:
        return None


def getDetailInfo(url):
    # 提取详情页信息
    try:
        res = requests.get(url, headers=headers, timeout=10)
        e = etree.HTML(res.text)
        title = ''.join(e.xpath('//div[@class="ProjetInfo_ms"][1]//tr[10]/td[2]/p/text()')).strip()
        item_info = e.xpath('//div[@class="ProjetInfo_ms"][2]')
        left_name = ''.join(item_info[0].xpath('.//tr[1]/td[2]/p/text()')).strip()
        right_name = ''.join(item_info[0].xpath('.//tr[1]/td[4]/p/text()')).strip()
        left_phone = ''.join(item_info[0].xpath('.//tr[3]/td[2]/text()')).strip()
        right_phone = ''.join(item_info[0].xpath('.//tr[3]/td[4]/text()')).strip()
        unit = ''.join(item_info[0].xpath('.//tr[10]/td[2]/p/text()')).strip()
        print(index, title, left_name, right_name, left_phone, right_phone, unit)
        sh1.append([url, title, left_name, left_phone, right_name, right_phone, unit])
    except Exception as e:
        logging.error(f'{index}, {url}\n')


if __name__ == "__main__":
    headers = {
        'Host': 'www.chictr.org.cn',
        'Cookie': 'acw_tc=76b20f6716772248712666095e059afc7339fe967f9a184f4ae2015e774aae; onlineusercount=1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.60 Safari/537.36',
    }
    f = open('url.txt', 'a+', encoding='utf-8')
    for page in range(1, 336):
        # time.sleep(1)
        while True:
            print(f'正在采集第 {page} 页...')
            getVerifyimagepage()  # 获取验证码
            verifycode = imgRecognition('img.jpg')  # 识别验证码
            result = getList(page, verifycode)
            # 如果有返回数据，退出循环，否则重新获取识别验证码，重新请求
            if result:
                break
    f.close()

    wb = openpyxl.Workbook()
    all_sheetnames = wb.sheetnames
    sh1 = wb[all_sheetnames[0]]
    sh1.append(['链接', '标题', '申请注册联系人', '申请注册联系人电话', '研究负责人', '研究负责人电话', '申请人所在单位'])
    url_list = []
    with open('url.txt', 'r') as fp:
        for line in fp.readlines():
            url_list.append(line.strip().split(',')[-1])
    for index, url in enumerate(url_list):
        getDetailInfo(url)
        # break
    wb.save('all_data.xlsx')
    print('DONE!')
